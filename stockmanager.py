import openpyxl
from datetime import datetime
from PySide2 import QtCore, QtWidgets, QtGui
from PySide2.QtUiTools import QUiLoader
from PySide2.QtCore import QFile
import sys
from PySide2.QtWidgets import QMainWindow, QWidget


class StockManager:
    def __init__(self):
        self.workbook = None
        self.p_sheet = None
        self.t_sheet = None
        self.m_sheet = None
        self.product_id = None
        self.type = None
        self.count = None

        self.selected = {}                    # product : 다미끼 , weight : 250 , color : #01, size : #3
        

    def open_excel(self):
        self.workbook  = openpyxl.load_workbook('jiggingpro.xlsx')
        # self.t_sheet = self.workbook.get_sheet_by_name('Transaction')
        # self.p_sheet = self.workbook.get_sheet_by_name('Product')
        self.t_sheet = self.workbook['Transaction']
        self.p_sheet = self.workbook['Product']
        self.m_sheet = self.workbook['Manufacturer']



    def _update_count(self, row, cnt):
        self.p_sheet['G' + str(row)] = self.p_sheet['G' + str(row)].value + cnt

    def _find_row_by_product_ID(self, id):
        ret = None
        for i in range(2, self.p_sheet.max_row+1):
            if self.p_sheet['A'+str(i)].value == id :
                ret = i
                break
            else :
                pass
        return ret


    def update_stock(self):
        get_product_ID = lambda row : self.t_sheet['C'+str(row)].value
        get_count = lambda row : self.t_sheet['D'+str(row)].value
        get_type = lambda row : self.t_sheet['B'+str(row)].value
        p_n = lambda t : 1 if t == 'buy' else -1

        for row in range(2, self.t_sheet.max_row+1):
            id = get_product_ID(row)
            type = get_type(row)
            cnt = get_count(row)
            cnt = cnt * p_n(type)

            p_row = self._find_row_by_product_ID(id)
            self._update_count(p_row, cnt)

    def save(self):
        now = datetime.now()
        fname = 'jiggingpro'+str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+str(now.minute)+str(now.second)+'.xlsx'
        self.workbook.save(fname)

    def get_manufacturer(self):
        manuf_cells = lambda colm: self.m_sheet[colm]
        cell_value = lambda cell : cell.value

        manufacturer_list = list(map(cell_value, manuf_cells('B')))
        return manufacturer_list[1:]         # 첫줄의 카테고리 이름을 제외 시킨다.

    def get_products(self, name):            # name 은 제조사 이름 ex)다미끼
        self.selected['manufacturer'] = name
        product_sheet = self.workbook[name]

        product_cells = lambda colm: product_sheet[colm]
        cell_value = lambda cell: cell.value

        product_set = set(map(cell_value, product_cells('B')))
        product_set.remove('NAME')
        return list(product_set)

    def get_weight(self, name):               # name 은 제조사 이름 ex)랜스롱
        self.selected['product'] = name

        p_name = self.selected['product']
        product_sheet = self.workbook[p_name]

        product_cells = lambda colm: product_sheet[colm]
        cell_value = lambda cell: cell.value

        product_set = set(map(cell_value, product_cells('C')))

        









        


        












class stock_window:
    def __init__(self, stock_manager):
        self.sm = stock_manager
        self.ui_file = QFile("product_select.ui")
        self.ui_file.open(QFile.ReadOnly)
        self.loader = QUiLoader()
        self.main_window = self.loader.load(self.ui_file)
        self.ui_file.close()
        self.main_window.show()

        # 콤보박스 객체 저장
        self.combobox_manufacturer = None
        self.combobox_product = None
        self.combobox_weight = None
        self.combobox_color = None
        self.combobox_size = None

        # 콤보박스 선택된 item Text
        self.combobox_manufacturer_text = None
        self.combobox_product_text = None
        self.combobox_weight_text = None
        self.combobox_color_text = None
        self.combobox_size_text = None

        # 제조사 콤보박스 객체 찾기
        cb = self.main_window.findChild(QtWidgets.QComboBox, "comboBox")
        self.combobox_manufacturer = cb
        # 제조사 콤보박스에 시그널 별로 슬랏 연결
        cb.activated.connect(self.activated_manufacturer)

        # 제품 콤보박스 객체 찾기
        cb2 = self.main_window.findChild(QtWidgets.QComboBox, "comboBox_2")
        self.combobox_product = cb2
        # 제품 콤보박스에 시그널 슬랏 연결
        cb2.activated.connect(self.activated_product)

        # 무게 콤보박스 객체 찾기
        cb3 = self.main_window.findChild(QtWidgets.QComboBox, "comboBox_3")
        self.combobox_weight = cb3
        # 제품 콤보박스에 시그널 슬랏 연결
        cb3.activated.connect(self.activated_weight)

        # 컬러 콤보박스 객체 찾기
        cb4 = self.main_window.findChild(QtWidgets.QComboBox, "comboBox_4")
        self.combobox_color = cb4
        # 제품 콤보박스에 시그널 슬랏 연결
        cb3.activated.connect(self.activated_color)

        self.combobox_add_manufacturer()



    def combobox_add_manufacturer(self):
        cb = self.combobox_manufacturer
        # 엑셀 메니저에서 제조사 리스트 가져오기
        mf_list = self.sm.get_manufacturer()
        # 제조사 리스트 콤보박스에 아이템 추가
        cb.addItems(mf_list)



    def activated_manufacturer(self):
        print("activated")
        # 현재 선택된 아이템 저장
        text = self.combobox_manufacturer.currentText()
        self.combobox_manufacturer_text = text
        print(text)

        # 선택된 제조사를 보고 그 제조사 제품을 추가
        self.combobox_add_product()

    def combobox_add_product(self):
        # 제조사에 해당하는 제품 리스트 가져오기
        name = self.combobox_manufacturer_text
        p_list = self.sm.get_products(name)
        print(p_list)

        # 제품 콤보박스에 제품리스트 추가
        cb2 = self.combobox_product
        cb2.addItems(p_list)

    def activated_product(self):
        print("activated")
        # 현재 선택된 아이템 저장
        text = self.combobox_product.currentText()
        self.combobox_product_text = text
        print(text)

        # 선택된 제조사를 보고 그 제조사 제품을 추가
        self.combobox_add_weight()

    def combobox_add_weight(self):
        name = self.combobox_product_text
        w_list = self.sm.get_weight(name)
        print(w_list)

        cb3 = self.combobox_weight
        cb3.addItems(w_list)


    def activated_weight(self):
        print("activated")
        # 현재 선택된 아이템 저장
        text = self.combobox_weight.currentText()
        self.combobox_weight_text = text
        print(text)

        # 선택된 제조사를 보고 그 제조사 제품을 추가
        self.combobox_add_color()

    def combobox_add_color(self):
        name = self.combobox_weight_text
        c_list = self.sm.get_color(name)
        print(c_list)

        cb4 = self.combobox_color
        cb4.addItems(c_list)

    def activated_color(self):
        print("activated")
        # 현재 선택된 아이템 저장
        text = self.combobox_color.currentText()
        self.combobox_color_text = text
        print(text)

        # 선택된 제조사를 보고 그 제조사 제품을 추가
        # self.combobox_add_size()

    def highlighted(self):
        print("highlighted")

    def idxchanged(self):
        print("idx changed")




if __name__ == '__main__':
    sm = StockManager()
    sm.open_excel()
    # sm.update_stock()
    # sm.save()
    app = QtWidgets.QApplication(sys.argv)
    sw = stock_window(sm)
    sys.exit(app.exec_())
    

